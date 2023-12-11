import os
import pathlib
import zipfile
from datetime import date, datetime
import pandas as pd
from pandas.api.types import is_datetime64_any_dtype as is_datetime

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from typing import Any, List, Tuple, Dict, Optional, Callable

import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

try:
    from plugins.com import excel
    from pywintypes import com_error
    from PIL import ImageGrab
    IMPORT_PYWIN32COM_OK = True
except ModuleNotFoundError:
    IMPORT_PYWIN32COM_OK = False


class ExcelRecipes:

    @classmethod
    def dataframe_to_excel(cls, 
                           df: pd.DataFrame, 
                           path: str, 
                           sheet_name: str, 
                           dt_as_str: bool = True,
                           on_sheet_completed: Callable[[Worksheet], None] = None):
        """
        Exporta DataFrame como excel usando openpyxl ao invés do pandas

        :param df: DataFrame com os dados para exportar
        :param path: caminho onde o excel será salvo
        :param sheet_name: nome da planilha
        :param dt_as_str: indica se converte colunas do tipo datetime para string
        """
        wb = openpyxl.Workbook()
        # pegamos a aba ativa (1ª por padrão)
        ws = wb.active
        # renomeamos o nome da aba
        ws.title = sheet_name

        # converte datetime para string, caso informado
        if dt_as_str:
            date_cols = [col for col in df.columns if is_datetime(df[col])]
            df[date_cols] = df[date_cols].fillna('').astype(str)

        # para cada linha do dataframe (pode-se pegar índice e cabeçalho) adiciona à aba
        for r in dataframe_to_rows(df.fillna(''), index=False, header=True):
            ws.append(r)

        # pega a coordenada final da planilha para criação da tabela
        end_coordinate = ws.cell(row=max(ws.max_row, 2), column=ws.max_column).coordinate

        tab = Table(displayName=f'Tabela1', ref=f'A1:{end_coordinate}')

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name='TableStyleMedium1', showFirstColumn=False,
                            showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # reajusta coluna a coluna de acordo para comportar o texto
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))+2))

        for col, value in dims.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = value

        if on_sheet_completed:
            on_sheet_completed(ws)

        # salva planilha
        wb.save(path)

    @classmethod
    def dataframes_to_excel(cls, 
                            path: str, 
                            sheets_data: List[Tuple[str, pd.DataFrame]], 
                            dt_as_str: bool = True,
                            on_sheet_completed: Callable[[Worksheet], None] = None):
        """
        Exporta vários DataFrame's como excel usando openpyxl ao invés do pandas

        :param path: caminho onde o excel será salvo
        :param sheets_data: lista de tuplas contendo nome da planilha e DataFrame
        :param dt_as_str: indica se converte colunas do tipo datetime para string
        """
        wb = openpyxl.Workbook()

        for i, sheet_data in enumerate(sheets_data):
            sheet_name, df = sheet_data

            if i == 0:
                # pegamos a aba ativa (1ª por padrão)
                ws = wb.active
                # renomeamos o nome da aba
                ws.title = sheet_name
            else:
                # pegamos a aba ativa (1ª por padrão)
                ws = wb.create_sheet(sheet_name)

            # converte datetime para string, caso informado
            if dt_as_str:
                date_cols = [col for col in df.columns if is_datetime(df[col])]
                df[date_cols] = df[date_cols].fillna('').astype(str)

            # para cada linha do dataframe (pode-se pegar índice e cabeçalho) adiciona à aba
            for r in dataframe_to_rows(df.fillna(''), index=False, header=True):
                ws.append(r)

            # pega a coordenada final da planilha para criação da tabela
            end_coordinate = ws.cell(row=max(ws.max_row, 2), column=ws.max_column).coordinate

            #tab = Table(displayName=sheet_name.replace(' ', ''), ref=f'A1:{end_coordinate}')
            tab = Table(displayName=f'Tabela{i+1}', ref=f'A1:{end_coordinate}')

            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(name='TableStyleMedium1', showFirstColumn=False,
                                showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)

            # reajusta coluna a coluna de acordo para comportar o texto
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))+2))

            for col, value in dims.items():
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = value

            if on_sheet_completed:
                on_sheet_completed(ws)

        # salva planilha
        wb.save(path)

    @classmethod
    def fill_template(cls, 
                      filename: str, 
                      save_as: str,
                      values_dict: Dict[str, Any], 
                      *, 
                      placeholder_pattern: Optional[str] = '{{placeholder}}'):
        """
        Preenche uma planilha substituindo termos por valores de um dicionário.

        :param filename: nome do arquivo de entrada
        :param save_as: nome do arquivo de saída
        :param values_dict: dicionário contendo os termos e os valores a serem substituídos
        :param placeholder_pattern: padrão do placeholder. Note que a palavra "placeholder" deve constar no padrão.
        """
        assert filename.lower().endswith('.xlsx'), 'Arquivo de entrada precisa ser uma planilha XLSX'
        assert save_as.lower().endswith('.xlsx'), 'Arquivo de saída precisa ser uma planilha XLSX'
        assert filename.lower()!=save_as.lower(), 'Os arquivos de entrada e saída não podem ser os mesmos.'
        assert 'placeholder' in placeholder_pattern, 'O padrão de substituição precisa conter a palavra "placeholder"'

        known_converters = {
            float: lambda v: str(v).replace('.', ','),
            date: lambda v: v.strftime(r'%d/%m/%Y'),
            datetime: lambda v: v.strftime(r'%d/%m/%Y'),
        }

        with zipfile.ZipFile(filename, 'r') as zin:
            with zipfile.ZipFile(save_as, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                zout.comment = zin.comment
                for item in zin.infolist():
                    # os arquivos que queremos modificar estão dentro da pasta xl
                    if os.path.dirname(item.filename)=='xl/worksheets' or \
                       item.filename=='xl/sharedStrings.xml' or \
                       item.filename=='xl/workbook.xml' or \
                       item.filename=='docProps/app.xml':
                        content = zin.read(item.filename).decode('utf-8')
                        for k, v in values_dict.items():
                            placeholder = placeholder_pattern.replace('placeholder', str(k))
                            converted_v = known_converters.get(type(v), str)(v)
                            content = content.replace(placeholder, converted_v)
                        zout.writestr(item.filename, content)
                    # do contrário, só copia
                    else:
                        zout.writestr(item.filename, zin.read(item.filename))

    @classmethod
    def fill_template_with_com(cls, filename: str, save_as: str, values_dict: Dict[str, Any], *, 
                               placeholder_pattern: Optional[str] = '{{placeholder}}'):
        """
        Preenche uma planilha substituindo termos por valores de um dicionário usando objetos COM.

        :param filename: nome do arquivo de entrada
        :param save_as: nome do arquivo de saída
        :param values_dict: dicionário contendo os termos e os valores a serem substituídos
        :param placeholder_pattern: padrão do placeholder. Note que a palavra "placeholder" deve constar no padrão.
        """
        assert IMPORT_PYWIN32COM_OK, 'Os requisitos para essa funcionalidade não foram instalados/ configurados.'

        # pega o caminho absoluto dos parâmetros
        filename = str(pathlib.Path(filename).resolve())
        save_as = str(pathlib.Path(save_as).resolve())

        assert filename.lower().endswith('.xlsx'), 'Arquivo de entrada precisa ser uma planilha XLSX'
        assert save_as.lower().endswith('.xlsx'), 'Arquivo de saída precisa ser uma planilha XLSX'
        assert filename.lower()!=save_as.lower(), 'Os arquivos de entrada e saída não podem ser os mesmos.'
        assert 'placeholder' in placeholder_pattern, 'O padrão de substituição precisa conter a palavra "placeholder"'

        known_converters = {
            float: lambda v: str(v).replace('.', ','),
            date: lambda v: v.strftime(r'%d/%m/%Y'),
            datetime: lambda v: v.strftime(r'%d/%m/%Y'),
        }

        with excel.application(kill_after=True) as xl:

            # abre planilha
            wb = xl.Workbooks.Open(filename)
            # guarda aba ativa para mantermos ela no final
            default_sheetname = wb.ActiveSheet.Name

            def update_worksheet_name(ws: Any):
                """ função interna para atualização do nome da aba, caso tenha placeholder """
                ws_name = str(ws.Name)

                for k, v in values_dict.items():
                    placeholder = placeholder_pattern.replace('placeholder', str(k))

                    if placeholder in ws_name:
                        converted_v = known_converters.get(type(v), str)(v)
                        ws.Name = excel.worksheet_name_sanitization(ws_name.replace(placeholder, converted_v))

            def update_cell(ws: Any, cell: Any):
                """ função interna para atualização do valor da célula """
                cell_value = str(cell.Value)

                for k, v in values_dict.items():
                    placeholder = placeholder_pattern.replace('placeholder', str(k))

                    if placeholder in cell_value:
                        # se a célula for o próprio placeholder e estamos falando de dados numéricos...
                        if placeholder == cell_value and isinstance(v, (int, float)):
                            cell.Value = v
                        # se a célula for o próprio placeholder e estamos falando de um dataframe...
                        elif placeholder == cell_value and isinstance(v, pd.DataFrame):
                            v.to_clipboard(excel=True, index=False, header=False, decimal=',', encoding='utf-8', date_format=r'%d/%m/%Y')
                            ws.Range(cell.Address).Select()
                            ws.Paste()
                        # do contrário vamos usar algum conversor para string
                        else:
                            v = v or ''
                            converted_v = known_converters.get(type(v), str)(v)
                            cell.Value = cell_value.replace(placeholder, converted_v)

            # a primeira rodada é para passarmos por todas as células
            for ws in wb.Worksheets:
                ws.Select()

                for c in ws.UsedRange.Cells:
                    update_cell(ws, c)

                ws.Range('A1').Select()

            # a segunda rodada é para passarmos pelo nome de todas as abas
            wb.Sheets(default_sheetname).Select()
            for ws in wb.Worksheets:
                update_worksheet_name(ws) 

            try:
                if os.path.exists(save_as):
                    os.remove(save_as)
                wb.SaveAs(save_as)
                wb.Close()
            except com_error:
                # as vezes ocorre erro, mas o arquivo foi salvo... então verificamos antes de acusar algum erro
                if not os.path.exists(save_as):
                    raise
            except:
                raise

    @classmethod
    def export_image(cls, excel_filename: str, image_filename: str, page: str = None, range_: str = None):
        """
        Exporta aba do Excel como imagem

        :param excel_filename: caminho do arquivo excel
        :param image_filename: caminho do arquivo de saída da imagem
        :param page: nome da aba que se quer extrair a imagem
        :param range_: range que se quer extrair a imagem
        """
        assert IMPORT_PYWIN32COM_OK, 'Os requisitos para essa funcionalidade não foram instalados/ configurados.'

        excel_filename = str(pathlib.Path(excel_filename).resolve())
        image_filename = str(pathlib.Path(image_filename).resolve())

        output_ext = os.path.splitext(image_filename)[-1].lower()
        if output_ext not in ('.gif', '.bmp', '.png'):
            raise ValueError(f'Unsupported image format: {output_ext}')
        
        image_format = output_ext.replace('.', '')

        # se ambas as informações forem passadas, converte em range
        if range_ is not None and page is not None and '!' not in range_:
            range_ = f"'{page}'!{range_}"

        with excel.application(kill_after=True) as xl:
            wb = xl.Workbooks.Open(excel_filename)

            if range_ is None:
                if page is None: page = 1
                try:
                    rng = wb.Sheets(page).UsedRange
                except com_error:
                    raise Exception(f'Failed locating used cell range on page {page}')
                except AttributeError:
                    # pode se tratar de uma aba só com gráficos
                    rng = wb.Sheets(page).Export(image_filename)
                    return
                if str(rng) == 'None':
                    # se não tiver um range indicado...
                    shapes = wb.Sheets(page).Shapes
                    if len(shapes) == 1:
                        rng = shapes[0]
                    else:
                        raise Exception(f'Failed locating used cells or single object to print on page {page}')
            else:
                try:
                    rng = wb.Application.Range(range_)
                except com_error:
                    raise Exception(f'Failed locating range {range_}')

            for _ in rng.parent.Shapes: pass

            retries, success = 100, False
            while not success:
                try:
                    rng.CopyPicture(excel.constants.xlScreen, excel.constants.xlBitmap)
                    im = ImageGrab.grabclipboard()
                    im.save(image_filename, image_format)
                    success = True
                except (com_error, AttributeError, OSError):
                    retries -= 1
                    if retries == 0: 
                        raise

            wb.Close()
        return

    @classmethod
    def to_pdf(cls, excel_filename: str, pdf_filename: str, sheetname: str = None):
        """
        Converte arquivo XLSX em PDF. Necessário ter o Excel instalado.

        :param excel_filename: nome do arquivo XLSX de entrada
        :param pdf_filename: nome do arquivo PDF de saída
        :param sheetname: nome da aba (quando não informado, usa a ativa)
        """
        assert IMPORT_PYWIN32COM_OK, 'Os requisitos para essa funcionalidade não foram instalados/ configurados.'

        with excel.application(kill_after=True) as xl:
            excel_filename = str(pathlib.Path(excel_filename).resolve())
            pdf_filename = str(pathlib.Path(pdf_filename).resolve())

            wb = xl.Workbooks.Open(excel_filename)

            if sheetname:
                ws = wb.Sheets(sheetname)
            else:
                ws = wb.ActiveSheet

            if os.path.exists(pdf_filename):
                os.remove(pdf_filename)

            ws.ExportAsFixedFormat(0, pdf_filename)
            wb.Close()
        return